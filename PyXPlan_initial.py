#!/usr/bin/env python3
# -*-coding:utf-8 -*

import os
import sys
#import json
import shutil
from pathlib import Path
import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox, simpledialog
from datetime import date, timedelta, datetime
import pyexcel as pe
import openpyxl as op
from openpyxl.cell import MergedCell
import arrow
from ics import Calendar, Event

############################################################################################################################################
################################## fonctions ###############################################################################################




#############################################################################################################################################
################################### classes


class Evt :
  # l'evennement a les attributs date, Heure de début, H de fin, description de l'evenement
  def __init__(self,dat=None,hdeb=None,hfin=None,description=None):
    self.jour=dat
    self.hd=hdeb
    self.hf=hfin
    self.desc=description

class PlanningPromo():
  # extrait le planning d'une promo depuis le fichier xlsm de FP
  # créer une fenêtre pour afficher le déroulé des actions
  def __init__(self,promo,fname,fen):
    self.fen_text=fen
    self.promoid=promo
    self.sfile=fname
    self.splan=op.load_workbook(fname,data_only=True) #charge le planning sans les formules (uniquement les valeurs affichées à l'écran)
    self.diclistevt={} #dictionnaire classé par n° de semaine de liste des événements qui seront complétées au fur et à mesure
    self.__colhoraires={3:["08:00","08:55"],4:["09:00","09:55"],
                        5:["10:15","11:10"],6:["11h15","12:10"],
                        7:["12:15","13:10"],8:["13:15","14:10"],
                        9:["14:15","15:10"],10:["15:15","16:10"],
                        11:["16:15","17:10"],12:["17:15","18:10"]
                       } # dictionnaire chaque colonne a une heure de début et de fin
    self.listevent()
  

  def annee(self):
    # renvoie l'année du planning FP
    idx=self.sfile.find('20')
    if idx!=-1 :#si '20' est trouvé, peut aller chercher l'année
      an=int(self.sfile[idx:idx+4])
      return an
    else :
      return date.today().year # si pas d'année trouvée dans le nom de fichier, renvoie l'année en cours
      
  def listevent(self):
    jours=["LUNDI","MARDI","MERCREDI","JEUDI","VENDREDI"]
    e=Evt()
    self.fen_text.insert(END,"Lecture du planning FP ...\n")

    for sh in self.splan.worksheets:
      nomsh=sh.title.upper()
      if "SEM" in nomsh:
        self.fen_text.insert(END,nomsh+"\n")
        listevt=[]
        nosem=sh.cell(row=1,column=2).value
        lines=range(3,sh.max_row) # lignes de 3 à nombre de lignes max
        for l in lines:
          c=1
          boolpromo=False
          while c<13: #s'arrête à la colonne 12
            valcel=sh.cell(row=l, column=c).value
            if c==1 : # si c'est la première colonne
              if valcel!=None : # si la cellule n'est pas vide
                if valcel in jours : # si la cellule est un jour de la semaine
                  datejour=sh.cell(row=l+1, column=c).value # récupère la date dans la cellule en dessous
                  e.jour=datejour
            elif c==2 : # si c'est la 2ème colonne, va chercher s'il trouve la promo
              if valcel!=None and isinstance(valcel,str) : #si la cellule n'est pas vide et est une chaîne de caractères
                if self.promoid in valcel.upper(): #s'il trouve la promo dans la valeur de la cellule (on met tout en majuscule)
                  boolpromo=True
                else :
                  break #stoppe la boucle while si pas la bonne promo
              else :
                break #stoppe la boucle while si pas de promo
            elif boolpromo==True : #ne passe à la suite que si la ligne correspond à la promo recherchée
              if valcel!=None :
                e.desc=valcel
                e.hd=self.__colhoraires[c][0] #récupère la valeur de début d'horaire en fonction de la colonne item 0 de la liste de la colonne
                if sh.cell(row=l, column=c).coordinate in sh.merged_cells : #si la cellule fait partie des cells fusionnées
                  for d in range(c+1,13):
                    cel=sh.cell(row=l, column=d)
                    if (cel.coordinate in sh.merged_cells) and (cel.value==None):
                      c=d
                    else :
                      break #arrête la boucle for d
                e.hf=self.__colhoraires[c][1]
                listevt.append(e)
##                print("ligne:",l,"colonne:",c)
                e=Evt() #remise à zéro de l'événemt une fois inséré dans la liste
                e.jour=datejour #on récupère la date du jour concerné
            c+=1
            
        if len(listevt)!=0: #si la liste des evenements de la semaine n'est pas nulle, ajoute la liste au dictionnaire des evenements
          self.diclistevt[nosem]=listevt

        
##    for key,liste in self.diclistevt.items():
##      print(str(key))
##      for ev in liste :
##        print(str(key),str(ev.jour),str(ev.desc),str(ev.hd),str(ev.hf))
##                  
                
class Dicweekevt():
  #permet d'extaire des données du dictionnaire des événements
  def __init__(self,dic):
    self.dic=dic

  def minsem(self):
    #renvoie la valeur minimale des clés du dictionnaire
    listcles=list()
    for k in self.dic.keys():
      listcles.append(k)
    return min(listcles)

  def maxsem(self):
    #renvoie la valeur maximale des clés du dictionnaire
    listcles=list()
    for k in self.dic.keys():
      listcles.append(k)
    return max(listcles)

class Dictics():
  # convertit le dictionnaire en un fichier ics
  def __init__(self,dic,promo):
    self.dic=dic
    self.minwk=int(Dicweekevt(self.dic).minsem())
    self.maxwk=int(Dicweekevt(self.dic).maxsem())
    self.promoid=promo

  def firstwk(self):
    #fait choisir la première semaine à extraire et la renvoie
    wk=simpledialog.askinteger(title="Première semaine",
                               prompt="Entrer la première semaine à extraire",
                               initialvalue=self.minwk,
                               maxvalue=self.maxwk,
                               minvalue=self.minwk)
    if isinstance(wk,int):
      return wk

  def lastwk(self):
    #fait choisir la dernière semaine à extraire et la renvoie
    wk=simpledialog.askinteger(title="Dernière semaine",
                               prompt="Entrer la dernière semaine à extraire",
                               initialvalue=self.maxwk,
                               minvalue=self.minwk,
                               maxvalue=self.maxwk)
    if isinstance(wk,int):
      return wk

  def formatics(self):
    #met en forme au format ics et renvoie un calendrier ics
    startwk=self.firstwk()
    endwk=self.lastwk()
    if startwk>endwk:
      messagebox.showinfo(title="Numéros de semaine inversés",
                          prompt="Le numéro de semaine de début est supérieur à celui de la fin. Ils seront donc inversés.")
      end=startwk
      startwk=endwk
      endwk=end
    c=Calendar()
    for key,evtlist in self.dic.items():
      if key in range(startwk,endwk+1):
        for evt in evtlist :
          e=Event()
          e.name=evt.desc
          a=arrow.get(evt.jour,tzinfo='local') #passe la date au format arrow et en heure locale
          e.begin=a.replace(hour=int(evt.hd[:2]),minute=int(evt.hd[3:]),tzinfo='local') #modifie les heures/min et les formatent en heures locales
##          print(e.begin)
          e.end=a.replace(hour=int(evt.hf[:2]),minute=int(evt.hf[3:]),tzinfo='local')
##          print(e.end)
          c.events.add(e) #ajoute l'événement au calendrier
    return c

  def filedest(self):
    #renvoie le nom du fichier ics pour sauvegarde du type "MCTA22B 33 52 2023-12-27.ics"
    nom=" ".join([self.promoid,
                  str(self.minwk),
                  str(self.maxwk),
                  str(date.today()),
                  ".ics"])
    return nom
    
#############################################################################################################################################
################################### programme

repDest = str(os.getcwd())  #affecte le dossier parent à 'repertoire'
promo = ""
fileName = ""

#instancie une fenêtre principale
root = Tk()
root.title("PyXPlan: Extraction du planning FP d'une promo et export en ics")
if sys.platform=="win32":
  root.iconbitmap("Biplan.ico")
  
root.geometry("800x600")
#crée une fenêtre text pour afficher le déroulement du programme
# et y ajoute une barre verticale de défilement
fen_text = Text(root)
fen_text.pack(side="left", padx=20)
scroll_y = Scrollbar(root, orient="vertical", command=fen_text.yview)
scroll_y.pack(side="left", expand=True, fill="y")
fen_text.configure(yscrollcommand=scroll_y.set)


while fileName == "":
  fileName = filedialog.askopenfilename(
      initialdir=repDest,
      title="Choisir le fichier de planning FP d'où extraire les données (xlsx ou xlsm)"
  )  # demande quel fichier extraire et l'affecte à fileName


while promo == "":
  question = simpledialog.askstring(
      title="Promo", prompt="Saisir la promo à extraire (ex : MCTA22A)")
  if question != None:
    promo = question.upper()  # formate la promo en capitales

texte=fileName+"\n"+promo+"\n"
               
fen_text.insert(END,texte)

pl=PlanningPromo(promo,fileName,fen_text) #extrait le planning
dictics=Dictics(pl.diclistevt,promo) #le formate en ics

dirname = filedialog.askdirectory(
    initialdir=repDest,
    title="Choisir le dossier où enregistrer les données extraites"
)  #demande de choisir un dossier de destination
if dirname != "":
  repDest = dirname  #modifie le répertoire en fonction du choix de l'utilisateur

os.chdir(repDest)
fileDestName=dictics.filedest()


with open(fileDestName,'w') as my_ics:
  # sauvegarde l'ics sous fileDestName
  my_ics.writelines(dictics.formatics().serialize_iter())

messagebox.showinfo(
    title="Fin de l'extraction",
    message="Extraction de la promo "+promo+" terminée.\n Appuyer sur ENTRÉE pour quitter.")
root.destroy()
