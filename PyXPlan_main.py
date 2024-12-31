#!/usr/bin/env python3
# -*-coding:utf-8 -*

import os
import sys
import json
import shutil
from pathlib import Path
import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox, simpledialog
import tkinter.scrolledtext as st
from datetime import date, timedelta, datetime
import pyexcel as pe
import openpyxl as op
from openpyxl.cell import MergedCell
import arrow
from ics import Calendar, Event
import datetime
import pytz
from xls2xlsx import XLS2XLSX
import re #pour faire des recherches dans les strings


############################################################################################################################################
################################## fonctions ###############################################################################################
def save_liste(liste,filename):
  with open (filename,"w") as fic:
    for item in liste:
      fic.write("%s\n" % item)
    
def date_min(cal):
  #renvoie la plus petite date d'un calendar
  lsorted=sorted(cal.events)
  return lsorted[0].begin.date()

def date_max(cal):
  #renvoie la plus grande date d'un calendar
  lsorted=sorted(cal.events)
  return lsorted[-1].end.date()

def dates_min_max(cala,calp):
  #renvoie la période commune à deux calendars
  mina=date_min(cala)
  minp=date_min(calp)
  maxa=date_max(cala)
  maxp=date_max(calp)
  return [max(mina,minp),min(maxa,maxp)]

def browseFileFP():
	#demande à l'utilisateur de choisir fichier excel de la planning FP
	global dic_dir
	repDest = dic_dir.dic["dirPlanFP"]  #affecte le dossier parent à 'repertoire'
	if repDest=="":
		repDest=os.getcwd()
	
	filefp = filedialog.askopenfilename(
          initialdir=repDest, title="Sélectionner le fichier Excel de la planification FP")
  
	if (filefp!=None) and (filefp!="") and (isinstance(filefp,str)):
		if "xls" in os.path.splitext(filefp)[1]: # vérifie que l'extension est bien un fichier excel
			dirfp=os.path.dirname(filefp)
			dic_dir.dic["dirPlanFP"]=dirfp
			dic_dir.savejson()
		else:
			msg=messagebox.showwarning(title="Mauvaise extension",message="Il faut choisir le fichier Excel téléchargé depuis Teams, dans les fichiers de la Programmation.")
			filefp = ""

	return filefp

    
    

def browseFileAurion():
  #demande à l'utilisateur de choisir le fichier aurion sur lequel travailler
  global dic_dir
  repDest = dic_dir.dic["dirPlanAurion"]
  if repDest=="":
    repDest=os.getcwd()

  filea = filedialog.askopenfilename(
    initialdir=repDest, title="Sélectionner le fichier ics extrait d'Aurion")
  
  if (filea!=None) and (filea!="") and (isinstance(filea,str)):
    if os.path.splitext(filea)[1]==".ics": # vérifie que l'extension est bien '.ics'
      dira=os.path.dirname(filea)
      dic_dir.dic["dirPlanAurion"]=dira
      dic_dir.savejson()
    else:
      msg=messagebox.showwarning(title="Mauvaise extension",message="Il faut choisir un fichier .ics téléchargé sur Aurion")
      filea = ""

  return filea

def choixPromo():
  #demande et renvoie le nom de la promo
  try:
    question = simpledialog.askstring(
      title="Promo", prompt="Saisir la promo à extraire (ex : MCTA22A)")
    if (question != None) and (question != ""):
      promo = question.upper()  # formate la promo en capitales
    else:
      promo = ""
      return
  except:
    err=erreur_log("Problème choix de promo",
                   "fonction : choixPromo")
    return
  return promo
  

def check_aurion_fp(text_box,filea,filep):
  #vérifie la cohérence des plannings Aurion et Fp et renvoie une liste des erreurs
  #crée un fichier ics consolidé des evts Aurion et FP pour avoir les salles et toutes les réunions
  global dic_dir
  caurion = Calendar()
  cpyxplan = Calendar()
  cal_consolide=Calendar()
  
##  repDest = dic_dir.dic["dirPlanAurion"]
##  if repDest=="":
##    repDest=os.getcwd()
##    
##  filea = filedialog.askopenfilename(
##      initialdir=repDest, title="Sélectionner le fichier ics exrait d'Aurion")
##  if filea!="":
##    dira=os.path.dirname(filea)
##    dic_dir.dic["dirPlanAurion"]=dira
##    dic_dir.savejson()
##    nomaurion=str(filea)[len(dira):]
##  else:
##    return

  try:
    with open(filea, encoding="utf-8", errors="ignore") as f:
      caurion = Calendar(f.read())
  except:
    err="Ouverture impossible de "+filea
    lieu="fonction check_aurion_fp ouverture 'filea'"
    erreur_log(err,lieu)
    return

  repDest = dic_dir.dic["dirPlanICS"]
  if repDest=="":
    repDest=os.getcwd()
    
  # filep = filedialog.askopenfilename(initialdir=repDest,
                                     # title="Sélectionner le fichier ics PyXPlan")

  if filep!="":
    dic_dir.dic["dirPlanICS"]=os.path.dirname(filep) #change le repertoire de dic_dir
    dic_dir.savejson() #et sauve
    try:
      with open(filep, encoding="utf-8", errors="ignore") as f:
        cpyxplan = Calendar(f.read())
    except:
      err="Ouverture impossible de "+filep
      lieu="fonction check_aurion_fp ouverture fichier ics PyXPlan : 'filep'"
      erreur_log(err,lieu)
      return
  
  #détermine la plage de dates commune des deux calendriers
  intercal=dates_min_max(caurion,cpyxplan)
  if intercal[0]>intercal[1]:
    #si les périodes ne correspondent pas, 
    #retourne un message d'erreur et sort de la fonction 
    messagebox.showerror(title="Incompatibilité plannings",
                         message="Les deux calendriers n\'ont pas de dates communes.")
    return
  
  listmatch = list()
  listmis = list()
  fmt="DD-MM-YYYY HH[h]mm" #format de date à renvoyer

  #extrait des calendars les evts utiles et les met sous forme de liste triée
  scalaurion,scalpyxplan,scalreunion=planning_sorted(caurion,cpyxplan,intercal[0],intercal[1])

  #parcourt d'abord le planing Aurion
  for ea in scalaurion:
    erreur=""
    astart = arrow.get(ea.begin.astimezone(pytz.timezone('utc')))
    match = False
    idesc=ea.description.upper().find("MATIÈRE")+len("MATIÈRE : ") #on cherche l'index de la matière dans la description
    ifindesc=ea.description.find("\n",idesc) #on cherche l'index de  fin de la matière
      
    for ep in scalpyxplan:
      pstart=ep.begin
      pend=ep.end

      if astart.date() == pstart.date():
        if astart.hour == pstart.hour:
          if astart.minute==pstart.minute:
            match = True
            #convertit les heures aurion en UTC
            utc_time=ea.begin.to("utc")
            ea.begin=utc_time
            utc_time=ea.end.to("utc")
            ea.end=utc_time
            #ajoute à la liste des événments consolidés
            listmatch.append(ea)
            break
          else:
            erreur="Heure de début incorrecte dans Aurion pour l'événement : "
            break
    if match == False:
      if erreur=="":
        erreur="Aucune correspondance trouvée dans le planning FP pour l'événement Aurion : "
      listmis.append(
        erreur+
        ","+
        astart.to("Europe/Paris").format(fmt) +
        " ," +
        str(ea.description[idesc:ifindesc])
        )
      text_box.insert(tk.INSERT,listmis[-1]+"\n") #affiche le dernier élt de la liste dans la fenêtre

  #simplifie le nom et enlève la description des evts
  for e in listmatch:
    sd=simpledesc(e)
    e.name=sd
    e.description=""
        
  #parcourt ensuite le planning FP
  for ep in scalpyxplan:

    pstart=arrow.get(ep.begin)
    pend=ep.end
    erreur=""
    match = False
    for ea in scalaurion:
      astart = ea.begin.astimezone(pytz.timezone('utc'))
      if astart.date() == pstart.date():
        if astart.hour == pstart.hour:
          if astart.minute==pstart.minute:
            match = True
            break
          else:
            erreur="Heure de début incorrecte dans Aurion pour l'événement : "
            break
    if match == False:
      if erreur=="":
        erreur="Aucune correspondance trouvée dans le planning Aurion pour l'événement FP : "
      listmis.append(
        erreur+
        ","+
        pstart.to('Europe/Paris').format(fmt) +
        " ," +
        str(ep.name)
        )
      #ajoute alors l'evt FP dans la liste consolidée pour que l'evt apparaisse ensuite tout de même dans l'ics
      listmatch.append(ep)
      text_box.insert(tk.END,listmis[-1]+"\n") #affiche le dernier élt de la liste dans la fenêtre
      text_box.see(tk.END) #pour afficher le défilement

  #ajoute ensuite toutes les réunions dans listmatch
  for e in scalreunion:
    listmatch.append(e)

  #sélectionne le répertoire d'export
  os.chdir(dic_dir.dic["dirExportPyXPlan"])
  
  #exporte listmatch en fichier ics
  for e in listmatch: # parcourt la liste des événements
    cal_consolide.events.add(e) #ajoute chaque evt au calendrier
  fileDestName="Planning consolidé "+intercal[0].strftime("%d_%m_%y")+" "+intercal[1].strftime("%d_%m_%y")+".ics"
  with open(fileDestName,'w') as my_ics:
    # sauvegarde l'ics sous fileDestName
    my_ics.writelines(cal_consolide.serialize_iter())
  messagebox.showinfo(
      title="Création terminée",
      message="Le calendrier consolidé a été créé.\n Appuyer sur ENTRÉE pour quitter.")

###################  
## pour debug    
##  save_liste(listmatch,"listematch.txt")
###################
  
  #traitement des incohérences
  if listmis==[]:
    text_box.insert(tk.END,"Vérification terminée. Pas de problème Aurion.")
    messagebox.showinfo(title="Vérification terminée",message="Aucune incohérence trouvée entre Aurion et FP.")    
  else :  
    save_liste(listmis,"Incoherence Aurion "+str(date.today())+".txt")
    messagebox.showinfo(title="Vérification terminée",message="Des incohérences ont été détectées.\nFichier des incohérences Aurion enregistré.")
  
  #print(listmatch, sep='\n')
  #print(listmis, sep="\n")
  

def fp_to_ics(fen_text,fpfilename="", prom="",weekchoice=False):
  #extrait du planning FP et enregistre en ics
  #dic_dir est une classe de dictionnaire de répertoire
  #fen_text doit être le widget text (ici ce sera text_details)
  #weekchoice : si True on peut choisir l'intervalle des semaines à extraire, si False on prend toutes les semaines du planning
  
  global dic_dir
  repDest = dic_dir.dic["dirExportPyXPlan"]  #affecte le dossier parent à 'repertoire'
  if repDest=="":
    repDest=dic_dir.get_dir_planfp()
    
  promo = prom
  fileName = fpfilename
  
  if fileName == "" :
    try:
      fileName = filedialog.askopenfilename(
        initialdir=repDest,
        title="Fichier EDT FP (xls ou xlsm)"
        )  # demande quel fichier extraire et l'affecte à fileName

      if (str(fileName).find(".xls")<0) or (str(fileName)==""):
        messagebox.showwarning(title="Problème classeur EDT",message="Il faut choisir le classeur Excel FP "+ str(fileName))
        return
      else:
        dic_dir.dic["dirPlanFP"]=os.path.dirname(fileName)
        dic_dir.savejson()
    except:
      err=erreur_log("Problème choix du fichier Excel FP dans le répertoire '"+repDest+"'",
                      "fonction : fp_to_ics")
      return

  if promo == "" :
    promo = choixPromo() 

  texte=fileName+"\n"+promo+"\n"

  fen_text.insert(tk.END,texte+"\n")

  pl=PlanningPromo(promo,fileName,fen_text) #extrait le planning
  dictics=Dictics(pl.diclistevt,promo) #le formate en ics

# obsolète en V4
##  dirname = filedialog.askdirectory(
##      initialdir=dic_dir.dic["dirPlanICS"],
##      title="Choisir le dossier où enregistrer les données extraites"
##  )  #demande de choisir un dossier de destination
##  if dirname != "":
##    repDest = dirname  #modifie le répertoire en fonction du choix de l'utilisateur

  icscal=dictics.formatics(weekchoice) #transforme en calendar ics
  os.chdir(repDest)
  fileDestName=dictics.filedest()

  with open(fileDestName,'w') as my_ics:
    # sauvegarde l'ics sous fileDestName
    my_ics.writelines(icscal.serialize_iter())
    #my_ics.writelines(line.encode() +b"\n" for line in dictics.formatics().serialize_iter())

# obsolète V4
##  messagebox.showinfo(
##      title="Fin de l'extraction",
##      message="Extraction de la promo "+promo+" terminée.\n Appuyer sur ENTRÉE pour quitter.")

  return os.path.join(repDest,fileDestName) #renvoie le nom complet du fichier ics pour pouvoir l'utiliser plus tard
  
def erreur_log(erreur,location):
  #écrit dans un fichier le rapport des erreurs
  with open("log.txt","w") as filog:
    filog.write(erreur+
               "dans "+
               location)
    messagebox.showerror(title="Erreur",
                         message="Le programme a rencontré une erreur.\nCelle-ci a été enregistrée dans un fichier log que tu peux envoyer à Olivier Ternisien."
                        )

def planning_sorted(calaurion,calpp,ddeb,dfin):
  #tri les planning pour le réduire à la FP  
  #entre les dates ddeb et dfin et ordre chronologique
  #renvoie une liste d'évts aurion, une liste FP élèves et une liste d'evts réunions
  #typ doit être "aurion" ou "fp" en fonction du calendrier 
  l_evts_aurion=list()
  l_evts_fp=list()
  l_evts_reunion=list()

  l_txt=["FORMATION PRATIQUE","TRAINING EVALUATION"]
  for e in sorted(calaurion.events) :
    if e.begin.date()>=ddeb and e.begin.date()<=dfin:
      for t in l_txt:
        if (t in str(e.name).upper()) or (t in str(e.description).upper()):
          #si le texte est trouvé
          e.begin=e.begin.replace(tzinfo='Europe/Paris')
          e.end=e.end.replace(tzinfo='Europe/Paris')
          l_evts_aurion.append(e)
          break

  l_txt=["BILAN","ASTRAL","HARMO"]
  for e in sorted(calpp.events):
    test=False
    if e.begin.date()>=ddeb and e.begin.date()<=dfin:
      for t in l_txt:
        if t in str(e.name).upper():
          test=True
          l_evts_reunion.append(e)
          #si le texte est trouvé, renvoie true et ajoute l'evt à la liste des réunions
      if test==False:
        l_evts_fp.append(e)
  ###################################################################################################"
                   #ne sert que pour débug

##  save_liste(l_evts_aurion,"aurion trié.txt") #ne sert que pour débug
##  save_liste(l_evts_fp,"fp trié.txt") #ne sert que pour débug
##  save_liste(l_evts_reunion,"reunions trié.txt") #ne sert que pour débug
#####################################################################################################      

  return l_evts_aurion,l_evts_fp,l_evts_reunion

def simpledesc(e):
  #simplifie la descritption d'un evenement ics en ne mettant que le groupe et la matière
  desc=e.description
  name=e.name
  i=len("mcta22c_s56_")
  groupe=""
  if name.find('GR')>-1 :
      lgroupe=re.findall('GR.',name) #récupère une liste de tous les  groupes dans le summary de l'evt
      for g in lgroupe:
        groupe+=g
  mat=desc
  if desc.find("Matière :")>-1:
      m=re.search("Matière :.*",desc,flags=re.IGNORECASE) #cherche la matière dans la description
      mat=m.group()[len("Matière :"):] #récupère le string après 'matière' : et le remplace dans mat'
  newname=groupe+mat
  return newname
  
  

#############################################################################################################################################
################################### classes



class Directories:
  #gère les dir pour lire et sauvegarder les fichiers et stocke dans un fichier json
  def __init__(self):
    self.main=str(os.getcwd()) # à l'initialisation affecte le répertoire du programme
    self.nofile=False
    self.dic=self.__lirejson()
    self.__direxport()

  def __direxport(self):
    #crée le dossier où exporter les aurion consolidés et logs d'erreurs éventuellement détectées
    #si le dossier existe déjà, ne fait rien
    nomrep=os.path.join(self.main,"Exports Aurion")
    os.makedirs(nomrep,exist_ok=True) #crée le dossier s'il n'existe pas
    self.dic["dirExportPyXPlan"]=nomrep #complète le dictionnaire des répertoires
    self.savejson()
    
  def __lirejson(self):
    #lie le fichier json s'il existe et renvoie un dictionnaire de répertoires
    os.chdir(self.main)
    data={
      "dirPlanFP": "",
      "dirPlanAurion": "",
      "dirPlanICS": "",
      "dirExportPyXPlan":""
    }
    try :
      with open("directories.json","r") as fdir:
        sdata=json.load(fdir)
      if isinstance(sdata,dict):
        for key,d in sdata.items():
          data[key]=d
    except:
      self.nofile=True
    for key,d in data.items(): #parcourt le dictionnaire récupéré
      #si le champ est vide le remplace par le répertoire de travail
      if d=="" :
        data[key]=self.main
    return data

  def savejson(self):
    os.chdir(self.main)
    try:
      with open("directories.json","w") as fdir:
        json.dump(self.dic,fdir,indent=2)
    except:
      err=erreur_log("impossible de sauvegarder le fichier json",
                     "classe : Directories, fonction : savejson")
  def get_dir_planfp(self):
    try:
      dirname = filedialog.askdirectory(
        initialdir=self.dic["dirPlanICS"],
        title="Choisir le dossier où enregistrer les données extraites en ICS"
      )  #demande de choisir un dossier de destination
      if dirname != "":
        self.dic["dirPlanICS"] = dirname  #modifie le répertoire en fonction du choix de l'utilisateur
    except:
      err=erreur_log("Problème lors du choix du dossier de destination du fichier ICS",
         "classe : Directories, fonction : get_dirplanfp")
    return self.dic["dirPlanICS"]

  def get_dir_planaurion(self):
    try:
      dirname = filedialog.askdirectory(
        initialdir=self.dic["dirPlanAurion"],
        title="Choisir le dossier où trouver le planning extrait d'Aurion"
      )  #demande de choisir un dossier de destination
      if dirname != "":
        self.dic["dirPlanAurion"] = dirname  #modifie le répertoire en fonction du choix de l'utilisateur
    except:
      err=erreur_log("Problème lors du choix du dossier du planning extrait d'Aurion",
         "classe : Directories, fonction : get_dir_planaurion")
    return self.dic["dirPlanAurion"]
    

class Evt :
  # l'evennement a les attributs date, Heure de début, H de fin, description de l'evenement
  def __init__(self,dat=None,hdeb=None,hfin=None,description=None):
    self.jour=dat
    self.hd=hdeb
    self.hf=hfin
    self.desc=description

class PlanningPromo():
  # extrait le planning d'une promo depuis le fichier xlsm ou xlsx de FP
  # utilise une fenetre scrolledText passé en paramètre pour afficher le déroulé des actions
  def __init__(self,promo,fname,fen):
    self.fen_text=fen
    self.promoid=promo
    self.sfile=fname  
    self.splan=self.open_wb()#charge le planning sans les formules (uniquement les valeurs affichées à l'écran)
    self.diclistevt={} #dictionnaire classé par n° de semaine de liste des événements qui seront complétées au fur et à mesure
    self.__colhoraires={3:["08:00","08:55"],4:["09:00","09:55"],
                        5:["10:15","11:10"],6:["11h15","12:10"],
                        7:["12:15","13:10"],8:["13:15","14:10"],
                        9:["14:15","15:10"],10:["15:15","16:10"],
                        11:["16:15","17:10"],12:["17:15","18:10"]
                       } # dictionnaire chaque colonne a une heure de début et de fin
    self.listevent()

  def open_wb(self):
    #convertit éventuellement en xlsx
    #puis ouvre et renvoie le wbk pour pouvoir etre utilisé
    #charge sans les formules (data_only)
    ixls=self.sfile.find(".xls")
    if ixls==(len(self.sfile)-4):
      #si l'extension est xls, convertit en xlsx
      xfile=self.sfile+"x" #ajoute un x à l'extension du nom de fichier
      print(xfile)
      x2x=XLS2XLSX(self.sfile) #convertit en xlsx
      x2x.to_xlsx(xfile) #sauve en xlsx
      self.sfile=xfile #attribut le nouveau nom à sfile
      
    return op.load_workbook(self.sfile,data_only=True)
      
    
  def annee(self):
    # renvoie l'année du planning FP
    idx=self.sfile.find('20')
    if idx!=-1 :#si '20' est trouvé, peut aller chercher l'année
      an=int(self.sfile[idx:idx+4])
      return an
    else :
      return date.today().year # si pas d'année trouvée dans le nom de fichier, renvoie l'année en cours

  def listevent(self):
    jours=["LUNDI","MARDI","MERCREDI","JEUDI","VENDREDI"]
    e=Evt()
    self.fen_text.insert(tk.END,"Lecture du planning FP ...\n")

    for sh in self.splan.worksheets:
      nomsh=sh.title.upper()
      if "SEM" in nomsh:
        self.fen_text.insert(tk.END,nomsh+" - ")
        listevt=[]
        nosem=sh.cell(row=1,column=2).value
        lines=range(3,sh.max_row) # lignes de 3 à nombre de lignes max
        for l in lines:
          c=1
          boolpromo=False
          while c<13: #s'arrête à la colonne 12
            valcel=sh.cell(row=l, column=c).value
            if c==1 : # si c'est la première colonne
              if valcel!=None : # si la cellule n'est pas vide
                if valcel in jours : # si la cellule est un jour de la semaine

                  #récupère la date dans l'une des cellules en-dessous
                  lignedate = l + 1
                  while sh.cell(row=l+1, column=c).value == None :
                    lignedate += 1
                    
                  datejour = sh.cell(row=lignedate, column=c).value # récupère la date dans la cellule
                  e.jour=datejour
                  
            elif c==2 : # si c'est la 2ème colonne, va chercher s'il trouve la promo
              if valcel!=None and isinstance(valcel,str) : #si la cellule n'est pas vide et est une chaîne de caractères
                if self.promoid in valcel.upper(): #s'il trouve la promo dans la valeur de la cellule (on met tout en majuscule)
                  boolpromo=True
                else :
                  break #stoppe la boucle while si pas la bonne promo
              else :
                break #stoppe la boucle while si pas de promo
            elif boolpromo==True : #ne passe à la suite que si la ligne correspond à la promo recherchée
              if valcel!=None :
                e.desc=valcel
                # enregistre la description de l'événement en utf8
                e.hd=self.__colhoraires[c][0] #récupère la valeur de début d'horaire en fonction de la colonne item 0 de la liste de la colonne
                if sh.cell(row=l, column=c).coordinate in sh.merged_cells : #si la cellule fait partie des cells fusionnées
                  for d in range(c+1,13):
                    cel=sh.cell(row=l, column=d)
                    if (cel.coordinate in sh.merged_cells) and (cel.value==None):
                      c=d
                    else :
                      break #arrête la boucle for d
                e.hf=self.__colhoraires[c][1]
                listevt.append(e)
##                print("ligne:",l,"colonne:",c)
                e=Evt() #remise à zéro de l'événemt une fois inséré dans la liste
                e.jour=datejour #on récupère la date du jour concerné
            c+=1

        if len(listevt)!=0: #si la liste des evenements de la semaine n'est pas nulle, ajoute la liste au dictionnaire des evenements
          self.diclistevt[nosem]=listevt

        # ajoute un saut de ligne dans la fenêtre texte à la fin du process  
        self.fen_text.insert(tk.END, "\n")

class Dicweekevt():
  #permet d'extaire des données du dictionnaire des événements
  def __init__(self,dic):
    self.dic=dic

  def minsem(self):
    #renvoie la valeur minimale des clés du dictionnaire
    listcles=list()
    for k in self.dic.keys():
      listcles.append(k)
    return min(listcles)

  def maxsem(self):
    #renvoie la valeur maximale des clés du dictionnaire
    listcles=list()
    for k in self.dic.keys():
      listcles.append(k)
    return max(listcles)

class Dictics():
  # convertit le dictionnaire en un fichier ics
  def __init__(self,dic,promo):
    self.dic=dic
    self.fwk=0
    self.lwk=0
    self.minwk=int(Dicweekevt(self.dic).minsem())
    self.maxwk=int(Dicweekevt(self.dic).maxsem())
    self.promoid=promo

  def firstwk(self):
    #fait choisir la première semaine à extraire et la renvoie
    wk=simpledialog.askinteger(title="Première semaine",
                               prompt="Entrer la première semaine à extraire",
                               initialvalue=self.minwk,
                               maxvalue=self.maxwk,
                               minvalue=self.minwk)
    if isinstance(wk,int):
      #self.fwk=wk
      return wk

  def lastwk(self):
    #fait choisir la dernière semaine à extraire et la renvoie
    wk=simpledialog.askinteger(title="Dernière semaine",
                               prompt="Entrer la dernière semaine à extraire",
                               initialvalue=self.maxwk,
                               minvalue=self.minwk,
                               maxvalue=self.maxwk)
    if isinstance(wk,int):
      #self.lwk=wk
      return wk

  def formatics(self, choixsemaine=False):
    #met en forme au format ics et renvoie un calendrier ics

    # si choix semaine est False, prend toutes les semaines du planning
    if choixsemaine == True :
      startwk=self.firstwk()
      endwk=self.lastwk()
    else :
      startwk = self.minwk
      endwk = self.maxwk
    
    if startwk>endwk:
      messagebox.showinfo(title="Numéros de semaine inversés",
                          prompt="Le numéro de semaine de début est supérieur à celui de la fin. Ils seront donc inversés.")
      end=startwk
      startwk=endwk
      endwk=end

      
    self.fwk=startwk
    self.lwk=endwk
    
    c=Calendar()
    for key,evtlist in self.dic.items():
      if key in range(startwk,endwk+1):
        for evt in evtlist :
          e=Event()
          e.name=evt.desc
          a=arrow.get(evt.jour,tzinfo='Europe/Paris') #passe la date au format arrow et en heure locale
          e.begin=a.replace(hour=int(evt.hd[:2]),minute=int(evt.hd[3:]),tzinfo='Europe/Paris') #modifie les heures/min et les formatent en heures locales
##          print(e.begin)
          e.end=a.replace(hour=int(evt.hf[:2]),minute=int(evt.hf[3:]),tzinfo='Europe/Paris')
##          print(e.end)
          c.events.add(e) #ajoute l'événement au calendrier
    return c

  def filedest(self):
    #renvoie le nom du fichier ics pour sauvegarde du type "MCTA22B 33 52 2023-12-27.ics"
    nom=" ".join([self.promoid,
                  str(self.fwk),
                  str(self.lwk),
                  str(date.today()),
                  ".ics"])
    return nom

#############################################################################################################################################
################################### programme ####################

dic_dir=Directories()


# si le prog est lancé en autonome
if __name__== '__main__' :
  window = tk.Tk()  #instancie une fenêtre
  window.title("PyXPlan : extraction et vérification des plannings FP")
  window.configure(bg="black")
  bgcolor="darkslategray4"
  # ajoute une icone
  icon_file="Biplan.png"
  ico=tk.PhotoImage(file=icon_file)
  window.iconphoto(True,ico)


  frame_stext = tk.Frame(
    master=window,
    relief=tk.RIDGE,
    borderwidth=5)

  txt_details = st.ScrolledText(
    master=frame_stext,
    width=50,
    height=25,
    bg=bgcolor)
  #pour ajouter du texte dans le scrolledtext, txt_details.insert(tk.INSERT," ")

  frame_ics = tk.Frame(master=window, relief=tk.RIDGE, borderwidth=5)

  btn_export_to_ics = tk.Button(
    master=frame_ics,
    text="Sélectionner le fichier Excel de la plannification FP",
    width=50,
    height=5,
    bg=bgcolor,
    command=lambda: fp_to_ics(txt_details)
  )

  frame_aurion = tk.Frame(master=window, relief=tk.RIDGE, borderwidth=5)
  btn_check_aurion = tk.Button(
    master=frame_aurion,
    text="Vérifier cohérence avec Planning Aurion des élèves\n et \nCréer fichier ics consolidé",
    width=50,
    height=5,
    bg=bgcolor,
    command=lambda: check_aurion_fp(txt_details)
  )



  #formatte la fenetre (3 lignes et 1 colonne)
  for i in range(2):  #formatte la fenetre (3 lignes et 1 colonne
    window.rowconfigure(i, weight=1, minsize=50)

  window.columnconfigure(0, weight=1, minsize=75)

  #place les frame et les widgets associés
  frame_ics.grid(row=0, column=0, padx=5, pady=5)

  btn_export_to_ics.pack()

  frame_aurion.grid(row=1, column=0, padx=5, pady=5)
  btn_check_aurion.pack()

  frame_stext.grid(row=2, column=0, padx=5, pady=5)
  txt_details.pack()
  txt_details.configure(
      #state="disabled"
  )  #pour ne pas pouvoir laisser le curseur dans la fenêtre (readonly). Attention aucun texte n'apparaît par la suite

  window.mainloop()
